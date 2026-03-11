import sys
import os
import json
import re
import unicodedata
from datetime import datetime
from collections import defaultdict

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({
        "success": False,
        "error": "Biblioteca openpyxl não instalada. Execute: python -m pip install openpyxl"
    }, ensure_ascii=False))
    sys.exit(1)


BLOCK_ALIASES = {
    "plano_negocios": [
        "PLANO DE NEGOCIOS",
        "PLANO NEGOCIOS",
        "JBP",
        "PLANO DE INTRODUCAO",
        "PLANO INTRODUCAO",
    ],
    "historico": [
        "HISTORICO",
    ],
    "objetivos": [
        "OBJETIVOS",
        "OBJETIVO",
        "OBJETIVO DE VOLUME",
        "OBJETIVO NORTEADOR",
    ],
    "descricao_investimento": [
        "DESCRICAO DO INVESTIMENTO",
        "DESCRIÇÃO DO INVESTIMENTO",
        "INVESTIMENTO",
    ],
    "contrapartidas": [
        "CONTRAPARTIDAS",
        "CONTRAPARTIDAS EM FUNCAO DO INVESTIMENTO",
        "CONTRAPARTIDAS EM FUNÇÃO DO INVESTIMENTO",
    ],
    "itens_foco": [
        "CONTRAPARTIDAS ITENS FOCO",
        "ITENS FOCO",
        "ITENS DE FOCO",
    ],
    "encartes_obrigatorios": [
        "ENCARTES OBRIGATORIOS",
        "ENCARTES OBRIGATÓRIOS",
        "ENCARTES",
    ],
}

HEADER_LABELS = {
    "cliente": ["CLIENTE"],
    "periodo": [
        "PERIODO",
        "PERIODO DO PLANO",
        "PERIODO DA INTRODUCAO",
        "PERIODO DE PAGAMENTO PLANO",
        "PERIODO DE ACOES PLANO",
    ],
    "numero_acordo": [
        "NUMERO DE ACORDO",
        "NÚMERO DE ACORDO",
        "ACORDO",
        "N DO ACORDO",
        "Nº DO ACORDO",
    ],
}


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"\s+", " ", text)
    return text.upper().strip()


def value_to_str(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_number(text):
    if text is None:
        return None
    s = str(text).upper().strip()
    s = s.replace("CX", "").replace("CXS", "").replace("R$", "").replace("%", "")
    s = s.replace(" ", "")
    # tenta formato brasileiro
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        # quando tiver apenas ponto, pode ser milhar ou decimal
        # regra simples da V1
        if s.count(".") > 1:
            s = s.replace(".", "")
        elif s.count(".") == 1 and len(s.split(".")[-1]) == 3:
            s = s.replace(".", "")
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def month_name_to_number(token):
    mapping = {
        "JAN": 1, "JANEIRO": 1,
        "FEV": 2, "FEVEREIRO": 2,
        "MAR": 3, "MARCO": 3, "MARÇO": 3,
        "ABR": 4, "ABRIL": 4,
        "MAI": 5, "MAIO": 5,
        "JUN": 6, "JUNHO": 6,
        "JUL": 7, "JULHO": 7,
        "AGO": 8, "AGOSTO": 8,
        "SET": 9, "SETEMBRO": 9,
        "OUT": 10, "OUTUBRO": 10,
        "NOV": 11, "NOVEMBRO": 11,
        "DEZ": 12, "DEZEMBRO": 12,
    }
    return mapping.get(normalize_text(token))


def normalize_period(text):
    """
    V1: tenta gerar um texto normalizado simples.
    Mantém o original e devolve uma string amigável para conferência.
    """
    if not text:
        return None

    raw = str(text).strip()
    t = normalize_text(raw)

    # casos simples: 01/25, 1/25, JAN/25, JANEIRO/25
    m = re.match(r"^(\d{1,2})/(\d{2,4})$", raw.strip())
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        return f"{year:04d}-{month:02d}"

    m = re.match(r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})$", t)
    if m:
        month = month_name_to_number(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        if month:
            return f"{year:04d}-{month:02d}"

    # intervalo ex: JANEIRO/25 A DEZEMBRO/25
    m = re.match(r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})\s*(A|ATE|À)\s*([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})$", t)
    if m:
        m1 = month_name_to_number(m.group(1))
        y1 = int(m.group(2))
        m2 = month_name_to_number(m.group(4))
        y2 = int(m.group(5))
        if y1 < 100:
            y1 += 2000
        if y2 < 100:
            y2 += 2000
        if m1 and m2:
            return f"{y1:04d}-{m1:02d} a {y2:04d}-{m2:02d}"

    return raw


def worksheet_to_grid(ws):
    non_empty = []
    for row in ws.iter_rows():
        row_values = []
        for cell in row:
            val = value_to_str(cell.value)
            row_values.append(val)
        non_empty.append(row_values)
    return non_empty


def find_block_headers(grid):
    """
    Busca os aliases por célula.
    Guarda linha/coluna de início.
    """
    found = []
    for r_idx, row in enumerate(grid):
        for c_idx, value in enumerate(row):
            norm = normalize_text(value)
            if not norm:
                continue

            for block_name, aliases in BLOCK_ALIASES.items():
                for alias in aliases:
                    alias_norm = normalize_text(alias)
                    if alias_norm and alias_norm in norm:
                        found.append({
                            "block": block_name,
                            "row": r_idx,
                            "col": c_idx,
                            "label_found": value
                        })
                        break
    # remove duplicados muito próximos
    unique = []
    seen = set()
    for item in found:
        key = (item["block"], item["row"])
        if key not in seen:
            seen.add(key)
            unique.append(item)
    unique.sort(key=lambda x: (x["row"], x["col"]))
    return unique


def build_row_ranges(headers, total_rows):
    """
    V1: delimita bloco por faixa de linhas.
    Começa na linha do título e termina na linha anterior ao próximo título.
    """
    ranges = []
    if not headers:
        return ranges

    for i, item in enumerate(headers):
        start = item["row"]
        end = total_rows - 1
        if i < len(headers) - 1:
            end = headers[i + 1]["row"] - 1
        ranges.append({
            "block": item["block"],
            "start_row": start,
            "end_row": end,
            "start_col": item["col"],
            "label_found": item["label_found"]
        })
    return ranges


def slice_block(grid, block_info):
    rows = grid[block_info["start_row"]: block_info["end_row"] + 1]
    block_rows = []
    for row in rows:
        cleaned = [value_to_str(v) for v in row]
        if any(v for v in cleaned):
            block_rows.append(cleaned)
    return block_rows


def find_label_value_pairs(block_rows):
    """
    Heurística simples:
    - se a linha tem 2 ou mais células preenchidas, tenta usar 1ª como chave e 2ª como valor
    """
    pairs = []
    for row in block_rows:
        filled = [(idx, val) for idx, val in enumerate(row) if value_to_str(val)]
        if len(filled) >= 2:
            key = filled[0][1]
            value = filled[1][1]
            pairs.append({
                "label": key,
                "value": value,
                "row_raw": [v for _, v in filled]
            })
    return pairs


def extract_header(grid):
    """
    Procura cliente, período e número do acordo em toda a aba.
    """
    result = {
        "cliente": None,
        "periodo_original": None,
        "periodo_normalizado": None,
        "numero_acordo": None
    }

    for r_idx, row in enumerate(grid):
        filled = [(i, value_to_str(v)) for i, v in enumerate(row) if value_to_str(v)]
        for c_idx, val in filled:
            norm = normalize_text(val)
            for field_name, labels in HEADER_LABELS.items():
                for label in labels:
                    if normalize_text(label) in norm:
                        # tenta pegar a célula à direita
                        candidate = None
                        if c_idx + 1 < len(row):
                            candidate = value_to_str(row[c_idx + 1])
                        # fallback: próxima célula preenchida da linha
                        if not candidate:
                            for idx2, val2 in filled:
                                if idx2 > c_idx:
                                    candidate = val2
                                    break
                        if candidate:
                            if field_name == "cliente" and not result["cliente"]:
                                result["cliente"] = candidate
                            elif field_name == "periodo" and not result["periodo_original"]:
                                result["periodo_original"] = candidate
                                result["periodo_normalizado"] = normalize_period(candidate)
                            elif field_name == "numero_acordo" and not result["numero_acordo"]:
                                result["numero_acordo"] = candidate
    return result


def extract_historico(block_rows):
    text_lines = []
    metrics = []

    for row in block_rows:
        filled = [value_to_str(v) for v in row if value_to_str(v)]
        if not filled:
            continue
        line = " | ".join(filled)
        text_lines.append(line)

        # Tenta encontrar algo do tipo "descricao -> valor"
        joined = " ".join(filled)
        number_match = re.search(r"(\d[\d\.\,]*)\s*(CX|CXS|CAIXAS|%|R\$)?", joined.upper())
        if number_match:
            val = parse_number(number_match.group(1))
            metrics.append({
                "descricao_original": joined,
                "valor_numerico": val,
                "unidade_detectada": number_match.group(2) if number_match.group(2) else None
            })

    return {
        "texto_bruto": text_lines,
        "metricas_detectadas": metrics
    }


def extract_objetivos(block_rows):
    pairs = find_label_value_pairs(block_rows)
    normalized = []
    for item in pairs:
        normalized.append({
            "campo_original": item["label"],
            "valor_original": item["value"],
            "valor_numerico": parse_number(item["value"]),
        })

    return {
        "atributos": normalized,
        "texto_bruto": [" | ".join([v for v in row if value_to_str(v)]) for row in block_rows if any(value_to_str(v) for v in row)]
    }


def extract_descricao_investimento(block_rows):
    lines = []
    for row in block_rows:
        filled = [value_to_str(v) for v in row if value_to_str(v)]
        if filled:
            lines.append(" | ".join(filled))
    return {
        "texto_bruto": lines
    }


def extract_contrapartidas(block_rows):
    pairs = find_label_value_pairs(block_rows)
    rows_out = []
    for item in pairs:
        rows_out.append({
            "descricao": item["label"],
            "valor": item["value"],
            "valor_numerico": parse_number(item["value"])
        })
    return {
        "itens": rows_out,
        "texto_bruto": [" | ".join([v for v in row if value_to_str(v)]) for row in block_rows if any(value_to_str(v) for v in row)]
    }


def extract_itens_foco(block_rows):
    lines = []
    items = []
    for row in block_rows:
        filled = [value_to_str(v) for v in row if value_to_str(v)]
        if not filled:
            continue
        lines.append(" | ".join(filled))
        if len(filled) >= 2:
            items.append({
                "produto": filled[0],
                "quantidade_texto": filled[1],
                "quantidade_numerica": parse_number(filled[1])
            })

    return {
        "itens": items,
        "texto_bruto": lines
    }


def extract_encartes(block_rows):
    lines = []
    for row in block_rows:
        filled = [value_to_str(v) for v in row if value_to_str(v)]
        if filled:
            lines.append(" | ".join(filled))
    return {
        "texto_bruto": lines
    }


def process_sheet(ws):
    grid = worksheet_to_grid(ws)
    headers = find_block_headers(grid)
    ranges = build_row_ranges(headers, len(grid))

    raw_blocks = []
    identified_blocks = []
    result = {
        "sheet_name": ws.title,
        "identified_blocks": [],
        "header": extract_header(grid),
        "historico": None,
        "objetivos": None,
        "descricao_investimento": None,
        "contrapartidas": None,
        "itens_foco": None,
        "encartes_obrigatorios": None,
        "raw_blocks": []
    }

    for block_info in ranges:
        block_rows = slice_block(grid, block_info)
        block_name = block_info["block"]
        identified_blocks.append(block_name)

        raw_blocks.append({
            "block": block_name,
            "label_found": block_info["label_found"],
            "start_row": block_info["start_row"] + 1,
            "end_row": block_info["end_row"] + 1,
            "rows": [
                [v for v in row if value_to_str(v)]
                for row in block_rows
                if any(value_to_str(v) for v in row)
            ]
        })

        if block_name == "historico":
            result["historico"] = extract_historico(block_rows)
        elif block_name == "objetivos":
            result["objetivos"] = extract_objetivos(block_rows)
        elif block_name == "descricao_investimento":
            result["descricao_investimento"] = extract_descricao_investimento(block_rows)
        elif block_name == "contrapartidas":
            result["contrapartidas"] = extract_contrapartidas(block_rows)
        elif block_name == "itens_foco":
            result["itens_foco"] = extract_itens_foco(block_rows)
        elif block_name == "encartes_obrigatorios":
            result["encartes_obrigatorios"] = extract_encartes(block_rows)

    result["identified_blocks"] = sorted(list(set(identified_blocks)))
    result["raw_blocks"] = raw_blocks

    return result


def main():
    try:
        if len(sys.argv) < 2:
            raise Exception("Caminho do arquivo não informado.")

        file_path = sys.argv[1]

        if not os.path.exists(file_path):
            raise Exception("Arquivo não encontrado.")

        wb = load_workbook(file_path, data_only=True)

        output = {
            "success": True,
            "file_name": os.path.basename(file_path),
            "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_sheets": len(wb.sheetnames),
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output["sheets"].append(process_sheet(ws))

        print(json.dumps(output, ensure_ascii=False))

    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }, ensure_ascii=False))


if __name__ == "__main__":
    main()