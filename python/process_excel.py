import sys
import os
import json
import re
import unicodedata
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({
        "success": False,
        "error": "Biblioteca openpyxl não instalada. Execute: python -m pip install openpyxl"
    }, ensure_ascii=False))
    sys.exit(1)


MAIN_BLOCK_ALIASES = {
    "plano_negocios": [
        "PLANO DE NEGOCIOS",
        "PLANO DE INTRODUCAO",
        "JBP",
    ],
    "historico": [
        "HISTORICO",
    ],
    "objetivos": [
        "OBJETIVOS",
    ],
    "descricao_investimento": [
        "DESCRICAO DO INVESTIMENTO",
        "DESCRIÇÃO DO INVESTIMENTO",
    ],
    "contrapartidas_itens_foco": [
        "CONTRAPARTIDAS - ITENS FOCO",
        "CONTRAPARTIDAS ITENS FOCO",
        "ITENS FOCO",
    ],
    "contrapartidas_acoes": [
        "CONTRAPARTIDAS - ACOES",
        "CONTRAPARTIDAS - AÇÕES",
    ],
    "contrapartidas": [
        "CONTRAPARTIDAS",
        "CONTRAPARTIDA",
        "CONTRAPARTIDAS EM FUNCAO DO INVESTIMENTO",
        "CONTRAPARTIDAS EM FUNÇÃO DO INVESTIMENTO",
    ],
    "encartes_sugestao": [
        "ENCARTES SUGESTAO",
        "ENCARTES SUGESTÃO",
    ],
    "encartes_obrigatorios": [
        "ENCARTES OBRIGATORIOS",
        "ENCARTES OBRIGATÓRIOS",
        "SUGESTAO DE ENCARTES",
        "SUGESTÃO DE ENCARTES",
    ],
    "cadastros_vinculados": [
        "CADASTROS VINCULADOS",
        "CADASTROS VINCULADOS | LIBERACOES",
        "CADASTROS VINCULADOS | LIBERAÇÕES",
    ],
    "situacao_liberacao": [
    "SUBSTITUICAO | LIBERACAO",
    "SUBSTITUIÇÃO | LIBERAÇÃO",
    "SITUACAO LIBERACAO",
    "SITUAÇÃO LIBERAÇÃO",
    "SITUACAO DA LIBERACAO",
    "SITUAÇÃO DA LIBERAÇÃO",
    ],
    "oportunidades_cadastros_liberacoes": [
        "OPORTUNIDADE DE CADASTROS | LIBERACOES",
        "OPORTUNIDADE DE CADASTROS | LIBERAÇÕES",
        "OPORTUNIDADES DE CADASTROS | LIBERACOES",
        "OPORTUNIDADES DE CADASTROS | LIBERAÇÕES",
        "OPORTUNIDADES DE CADATROS | LIBERACOES",
        "OPORTUNIDADES DE CADATROS | LIBERAÇÕES",
    ],
    "investimentos_extras": [
        "INVESTIMENTOS EXTRAS",
    ],
}

FORM_TITLE_ALIASES = [
    "JBP",
    "PLANO DE NEGOCIOS",
    "PLANO DE INTRODUCAO",
]

HEADER_LABEL_ALIASES = {
    "cliente": [
        "CLIENTE",
        "CLIENTE:",
    ],
    "periodo_plano": [
        "PERIODO DO PLANO",
        "PERÍODO DO PLANO",
        "PERIODO DE PAGAMENTO PLANO",
        "PERÍODO DE PAGAMENTO PLANO",
    ],
    "periodo_acoes_plano": [
        "PERIODO DE ACOES PLANO",
        "PERÍODO DE AÇÕES PLANO",
    ],
    "numero_acordo": [
        "N DO ACORDO",
        "Nº DO ACORDO",
        "NUMERO DO ACORDO",
        "NÚMERO DO ACORDO",
        "N DO ACORDO:",
        "Nº DO ACORDO:",
    ],
}

CONTRAPARTIDAS_HEADERS = {
    "QUANTIDADE": "quantidade",
    "TIPO DE ACAO": "tipo_acao",
    "TIPO DE AÇÃO": "tipo_acao",
    "PERIODO": "periodo",
    "PERÍODO": "periodo",
    "LOJAS": "lojas",
    "OBS": "obs",
}

ITENS_FOCO_HEADERS = {
    "PRODUTO": "produto",
    "VOLUME MENSAL": "volume_mensal",
    "VOLUMEMENSAL": "volume_mensal",
    "VOLUME PERIODO": "volume_periodo",
    "VOLUME PERÍODO": "volume_periodo",
}

CADASTROS_HEADERS = {
    "PRODUTO": "produto",
    "LITRAGEM": "litragem",
    "VERSAO": "versao",
    "VERSÃO": "versao",
    "LOJAS": "abrangencia",
    "ABRANGENCIA": "abrangencia",
    "ABRANGÊNCIA": "abrangencia",
    "PRESENCA ATUAL": "abrangencia",
    "PRESENÇA ATUAL": "abrangencia",
}

MONTH_ALIASES = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}


def normalize_text(value):
    if value is None:
        return ""
    text = str(value).strip()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = re.sub(r"\s+", " ", text)
    return text.upper().strip()


def value_to_str(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    text = text.replace("\xa0", " ")
    return text.strip()


def cell_value_to_str(cell):
    value = cell.value
    if value is None:
        return ""

    if isinstance(value, (int, float)):
        fmt = str(getattr(cell, "number_format", "") or "")
        if "%" in fmt:
            percent_value = float(value) * 100
            decimals = 0
            percent_fmt = fmt.split("%", 1)[0]
            m = re.search(r"[.,](0+)$", percent_fmt)
            if m:
                decimals = len(m.group(1))
            formatted = f"{percent_value:,.{decimals}f}"
            formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
            return formatted + "%"

    return value_to_str(value)


def clean_scalar_text(value):
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\n+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def remove_empty_fields(obj):
    if isinstance(obj, dict):
        cleaned = {}
        for k, v in obj.items():
            vv = remove_empty_fields(v)
            if vv in ("", None, [], {}):
                continue
            cleaned[k] = vv
        return cleaned

    if isinstance(obj, list):
        cleaned_list = []
        for item in obj:
            vv = remove_empty_fields(item)
            if vv in ("", None, [], {}):
                continue
            cleaned_list.append(vv)
        return cleaned_list

    return obj


def excel_col_letter(col_idx_zero_based):
    col_num = col_idx_zero_based + 1
    result = ""
    while col_num:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


def parse_number(text):
    if text is None:
        return None

    raw = clean_scalar_text(text)
    if not raw:
        return None

    s = raw.upper()
    s = s.replace("R$", "")
    s = s.replace("%", "")
    s = s.replace("CAIXAS", "")
    s = s.replace("CAIXA", "")
    s = s.replace("CXS", "")
    s = s.replace("CX", "")
    s = s.replace(" ", "")

    if not s:
        return None

    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    else:
        if s.count(".") > 1:
            s = s.replace(".", "")
        elif s.count(".") == 1:
            right = s.split(".")[-1]
            if len(right) == 3:
                s = s.replace(".", "")
        s = s.replace(",", ".")

    try:
        return float(s)
    except Exception:
        return None


def detect_unit(text):
    if not text:
        return None
    raw = str(text).upper()
    norm = normalize_text(text)

    if "R$" in raw:
        return "BRL"
    if "%" in raw:
        return "%"
    if "CAIXAS" in norm or "CAIXA" in norm or "CXS" in norm or "CX" in norm:
        return "CX"
    return None


def month_name_to_number(token):
    return MONTH_ALIASES.get(normalize_text(token))


def split_products_from_cell(text):
    if not text:
        return []

    s = str(text).replace("\r", "\n")
    parts = re.split(r"\n|[|]", s)
    out = []
    for p in parts:
        cleaned = p.strip()
        if cleaned:
            out.append(cleaned)
    return out


def normalize_period(raw_text):
    if not raw_text:
        return {
            "periodo_original": None,
            "periodo_inicio": None,
            "periodo_fim": None,
            "periodo_normalizado": None,
            "periodo_status": "ausente"
        }

    raw = clean_scalar_text(raw_text)
    t = normalize_text(raw)

    m = re.match(r"^(\d{1,2})\s*/\s*(\d{2,4})$", raw)
    if m:
        month = int(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        p = f"{year:04d}-{month:02d}"
        return {
            "periodo_original": raw,
            "periodo_inicio": p,
            "periodo_fim": p,
            "periodo_normalizado": p,
            "periodo_status": "normalizado"
        }

    m = re.match(r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})$", t)
    if m:
        month = month_name_to_number(m.group(1))
        year = int(m.group(2))
        if year < 100:
            year += 2000
        if month:
            p = f"{year:04d}-{month:02d}"
            return {
                "periodo_original": raw,
                "periodo_inicio": p,
                "periodo_fim": p,
                "periodo_normalizado": p,
                "periodo_status": "normalizado"
            }

    m = re.match(
        r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})\s*(A|ATE|À)\s*([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})$",
        t
    )
    if m:
        m1 = month_name_to_number(m.group(1))
        y1 = int(m.group(2))
        m2 = month_name_to_number(m.group(4))
        y2 = int(m.group(5))
        if y1 < 100:
            y1 += 2000
        if y2 < 100:
            y2 += 2000
        if m1 and m2:
            start = f"{y1:04d}-{m1:02d}"
            end = f"{y2:04d}-{m2:02d}"
            return {
                "periodo_original": raw,
                "periodo_inicio": start,
                "periodo_fim": end,
                "periodo_normalizado": f"{start} a {end}",
                "periodo_status": "normalizado"
            }

    m = re.match(
        r"^([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*(A|ATE|À)\s*([A-ZÇÃÕÉÊÁÍÓÚ]+)\s*/\s*(\d{2,4})$",
        t
    )
    if m:
        m1 = month_name_to_number(m.group(1))
        m2 = month_name_to_number(m.group(3))
        y = int(m.group(4))
        if y < 100:
            y += 2000
        if m1 and m2:
            start = f"{y:04d}-{m1:02d}"
            end = f"{y:04d}-{m2:02d}"
            return {
                "periodo_original": raw,
                "periodo_inicio": start,
                "periodo_fim": end,
                "periodo_normalizado": f"{start} a {end}",
                "periodo_status": "normalizado"
            }

    return {
        "periodo_original": raw,
        "periodo_inicio": None,
        "periodo_fim": None,
        "periodo_normalizado": raw,
        "periodo_status": "nao_interpretado"
    }


def worksheet_to_grid(ws):
    grid = []
    max_col = ws.max_column or 0

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if ws.row_dimensions[row_idx].hidden:
            grid.append([""] * max_col)
            continue

        grid.append([cell_value_to_str(cell) for cell in row])
    return grid


def non_empty_cells(row):
    return [(i, value_to_str(v)) for i, v in enumerate(row) if value_to_str(v)]


def detect_column_regions(grid, min_non_empty_rows=3, gap_tolerance=2, min_region_width=4):
    if not grid:
        return []

    max_cols = max((len(r) for r in grid), default=0)
    col_scores = []

    for c in range(max_cols):
        count = 0
        for row in grid:
            if c < len(row) and value_to_str(row[c]):
                count += 1
        col_scores.append(count)

    candidate_cols = [i for i, score in enumerate(col_scores) if score >= min_non_empty_rows]
    if not candidate_cols:
        return []

    groups = []
    start = candidate_cols[0]
    prev = candidate_cols[0]

    for col in candidate_cols[1:]:
        if col - prev <= gap_tolerance:
            prev = col
        else:
            groups.append((start, prev))
            start = col
            prev = col
    groups.append((start, prev))

    filtered = []
    for start_col, end_col in groups:
        width = end_col - start_col + 1
        if width >= min_region_width:
            filtered.append({
                "start_col": start_col,
                "end_col": end_col,
                "width": width
            })

    return filtered


def crop_grid_to_region(grid, start_col, end_col):
    cropped = []
    for row in grid:
        if start_col < len(row):
            cropped.append(row[start_col:end_col + 1])
        else:
            cropped.append([])
    return cropped


def find_form_anchors(grid, max_header_rows=20):
    anchors = []

    max_rows = min(len(grid), max_header_rows)
    for r_idx in range(max_rows):
        row = grid[r_idx]
        for c_idx, value in enumerate(row):
            norm = normalize_text(value)
            if not norm:
                continue

            for alias in FORM_TITLE_ALIASES:
                alias_norm = normalize_text(alias)
                if norm == alias_norm or norm.startswith(alias_norm + " |") or norm.startswith(alias_norm + " -"):
                    anchors.append({
                        "row": r_idx,
                        "col": c_idx,
                        "label": value_to_str(value)
                    })
                    break

    anchors.sort(key=lambda x: x["col"])

    deduped = []
    for item in anchors:
        if not deduped:
            deduped.append(item)
            continue

        prev = deduped[-1]
        if abs(prev["col"] - item["col"]) <= 2:
            continue
        deduped.append(item)

    return deduped


def detect_form_regions_from_anchors(grid, anchors, side_margin=2):
    if not anchors:
        return []

    max_cols = max((len(r) for r in grid), default=0)
    regions = []

    for i, anchor in enumerate(anchors):
        if i == 0:
            start_col = max(0, anchor["col"] - side_margin)
        else:
            prev = anchors[i - 1]
            midpoint = (prev["col"] + anchor["col"]) // 2
            start_col = max(0, midpoint)

        if i == len(anchors) - 1:
            end_col = max_cols - 1
        else:
            nxt = anchors[i + 1]
            midpoint = (anchor["col"] + nxt["col"]) // 2
            end_col = max(start_col, midpoint - 1)

        regions.append({
            "start_col": start_col,
            "end_col": end_col,
            "width": end_col - start_col + 1,
            "anchor_col": anchor["col"],
            "anchor_label": anchor["label"]
        })

    return regions


def find_block_headers(grid):
    found = []

    for r_idx, row in enumerate(grid):
        for c_idx, value in enumerate(row):
            raw = value_to_str(value)
            norm = normalize_text(raw)
            if not norm:
                continue

            for block_name, aliases in MAIN_BLOCK_ALIASES.items():
                matched = False
                for alias in aliases:
                    alias_norm = normalize_text(alias)

                    is_match = norm == alias_norm
                    if not is_match:
                        is_match = (
                            norm.startswith(alias_norm + " |") or
                            norm.startswith(alias_norm + " -") or
                            norm.startswith(alias_norm + ":")
                        )
                    if not is_match:
                        is_match = norm.startswith(alias_norm + " ")

                    if is_match:
                        found.append({
                            "block": block_name,
                            "row": r_idx,
                            "col": c_idx,
                            "label_found": raw
                        })
                        matched = True
                        break

                if matched:
                    break

    found.sort(key=lambda x: (x["row"], x["col"]))

    deduped = []
    for item in found:
        if not deduped:
            deduped.append(item)
            continue

        prev = deduped[-1]
        if prev["block"] == item["block"] and abs(prev["row"] - item["row"]) <= 1:
            if len(item["label_found"]) > len(prev["label_found"]):
                deduped[-1] = item
        else:
            deduped.append(item)

    return deduped


def is_valid_form_region(headers_found):
    blocks = {h["block"] for h in headers_found}

    if "plano_negocios" not in blocks:
        return False

    meaningful = {
        "plano_negocios",
        "historico",
        "objetivos",
        "descricao_investimento",
        "contrapartidas",
        "contrapartidas_itens_foco",
        "contrapartidas_acoes",
        "encartes_obrigatorios",
        "encartes_sugestao",
        "cadastros_vinculados",
        "situacao_liberacao",
        "oportunidades_cadastros_liberacoes",
        "investimentos_extras",
    }

    found_meaningful = blocks.intersection(meaningful)
    return len(found_meaningful) >= 2


def build_row_ranges(headers, total_rows):
    if not headers:
        return []

    ranges = []
    for i, item in enumerate(headers):
        start = item["row"]
        end = total_rows - 1
        if i < len(headers) - 1:
            end = headers[i + 1]["row"] - 1

        ranges.append({
            "block": item["block"],
            "label_found": item["label_found"],
            "start_row": item["row"],
            "end_row": end,
            "start_col": item["col"]
        })
    return ranges


def slice_block_rows(grid, block_info):
    rows = []
    for r_idx in range(block_info["start_row"], block_info["end_row"] + 1):
        if r_idx >= len(grid):
            continue
        row = grid[r_idx]
        filled = non_empty_cells(row)
        if filled:
            rows.append({
                "row_excel": r_idx + 1,
                "cells": [
                    {
                        "col_idx": idx,
                        "col_excel": excel_col_letter(idx),
                        "cell_ref": f"{excel_col_letter(idx)}{r_idx + 1}",
                        "value": val
                    }
                    for idx, val in filled
                ]
            })
    return rows


def row_to_joined(row_obj):
    return " | ".join(clean_scalar_text(cell["value"]) for cell in row_obj["cells"])


def build_context(header, sheet_name, form_index):
    return {
        "cliente": header.get("cliente"),
        "periodo_original": header.get("periodo_original"),
        "periodo_inicio": header.get("periodo_inicio"),
        "periodo_fim": header.get("periodo_fim"),
        "periodo_normalizado": header.get("periodo_normalizado"),
        "periodo_status": header.get("periodo_status"),
        "periodo_acoes_original": header.get("periodo_acoes_original"),
        "periodo_acoes_inicio": header.get("periodo_acoes_inicio"),
        "periodo_acoes_fim": header.get("periodo_acoes_fim"),
        "periodo_acoes_normalizado": header.get("periodo_acoes_normalizado"),
        "periodo_acoes_status": header.get("periodo_acoes_status"),
        "numero_acordo": header.get("numero_acordo"),
        "titulo_plano": header.get("titulo_plano"),
        "aba_origem": sheet_name,
        "formulario_index": form_index
    }


def find_label_in_row(cells, aliases):
    aliases_norm = {normalize_text(a).rstrip(":") for a in aliases}
    for pos, cell in enumerate(cells):
        norm = normalize_text(clean_scalar_text(cell["value"])).rstrip(":")
        if norm in aliases_norm:
            return pos
    return None


def value_after_label(cells, label_pos):
    if label_pos is None:
        return None
    if label_pos + 1 < len(cells):
        return clean_scalar_text(cells[label_pos + 1]["value"])
    return None


def extract_header_from_block(rows_raw, label_found):
    header = {
        "cliente": None,
        "periodo_original": None,
        "periodo_inicio": None,
        "periodo_fim": None,
        "periodo_normalizado": None,
        "periodo_status": None,
        "periodo_acoes_original": None,
        "periodo_acoes_inicio": None,
        "periodo_acoes_fim": None,
        "periodo_acoes_normalizado": None,
        "periodo_acoes_status": None,
        "numero_acordo": None,
        "titulo_plano": clean_scalar_text(label_found)
    }

    for row in rows_raw:
        cells = row["cells"]

        pos_cliente = find_label_in_row(cells, HEADER_LABEL_ALIASES["cliente"])
        if pos_cliente is not None and not header["cliente"]:
            header["cliente"] = value_after_label(cells, pos_cliente)

        pos_periodo = find_label_in_row(cells, HEADER_LABEL_ALIASES["periodo_plano"])
        if pos_periodo is not None and not header["periodo_original"]:
            v = value_after_label(cells, pos_periodo)
            p = normalize_period(v)
            header["periodo_original"] = p["periodo_original"]
            header["periodo_inicio"] = p["periodo_inicio"]
            header["periodo_fim"] = p["periodo_fim"]
            header["periodo_normalizado"] = p["periodo_normalizado"]
            header["periodo_status"] = p["periodo_status"]

        pos_periodo_acoes = find_label_in_row(cells, HEADER_LABEL_ALIASES["periodo_acoes_plano"])
        if pos_periodo_acoes is not None and not header["periodo_acoes_original"]:
            v = value_after_label(cells, pos_periodo_acoes)
            p = normalize_period(v)
            header["periodo_acoes_original"] = p["periodo_original"]
            header["periodo_acoes_inicio"] = p["periodo_inicio"]
            header["periodo_acoes_fim"] = p["periodo_fim"]
            header["periodo_acoes_normalizado"] = p["periodo_normalizado"]
            header["periodo_acoes_status"] = p["periodo_status"]

        pos_acordo = find_label_in_row(cells, HEADER_LABEL_ALIASES["numero_acordo"])
        if pos_acordo is not None and not header["numero_acordo"]:
            header["numero_acordo"] = value_after_label(cells, pos_acordo)

    return header


def extract_year_from_title(title):
    if not title:
        return None
    m = re.search(r"(20\d{2}|\b\d{2}\b)", normalize_text(title))
    if not m:
        return None
    y = int(m.group(1))
    if y < 100:
        y += 2000
    return y


def parse_kv_list(rows_raw, header, sheet_name, form_index, tipo_registro, title_year=None):
    rows = []
    ctx = build_context(header, sheet_name, form_index)

    for order, row in enumerate(rows_raw, start=1):
        cells = list(row["cells"])
        joined = row_to_joined(row)

        label = ""
        value = ""

        while len(cells) >= 2:
            tail_label = clean_scalar_text(cells[-2]["value"])
            tail_value = clean_scalar_text(cells[-1]["value"])
            norm_tail_label = normalize_text(tail_label)
            norm_tail_value = normalize_text(tail_value)

            if re.match(r"^[A-Z]{3,4}-[A-Z]{3,4}$", norm_tail_label) and re.match(r"^\d{4}-\d{2}-\d{2}", norm_tail_value):
                cells = cells[:-2]
                continue

            break

        if len(cells) >= 2:
            label = clean_scalar_text(cells[0]["value"])
            value = clean_scalar_text(cells[-1]["value"])
        elif len(cells) == 1:
            single = clean_scalar_text(cells[0]["value"])
            if ":" in single:
                parts = single.split(":", 1)
                label = clean_scalar_text(parts[0])
                value = clean_scalar_text(parts[1])
            else:
                label = single
                value = ""
        else:
            continue

        norm_label = normalize_text(label)
        norm_value = normalize_text(value)
        if norm_label in {
            "OBJETIVOS",
            "DESCRICAO DO INVESTIMENTO",
            "DESCRIÇÃO DO INVESTIMENTO",
            "HISTORICO",
            "INVESTIMENTOS EXTRAS",
        }:
            continue

        if re.match(r"^[A-Z]{3,4}-[A-Z]{3,4}$", norm_label) and re.match(r"^\d{4}-\d{2}-\d{2}", norm_value):
            continue

        item = {
            **ctx,
            "tipo_registro": tipo_registro,
            "descricao": label,
            "valor_original": value,
            "valor_numerico": parse_number(value),
            "unidade": detect_unit(f"{label} {value}"),
            "ano_bloco": title_year,
            "linha_ordem": order,
            "row_excel": row["row_excel"],
            "linha_original": joined
        }
        rows.append(remove_empty_fields(item))

    return rows


def detect_table_header_row(rows_raw, expected_headers):
    expected_norm = set(normalize_text(x) for x in expected_headers.keys())
    best = None
    best_score = -1

    for idx, row in enumerate(rows_raw):
        row_norms = [normalize_text(clean_scalar_text(c["value"])) for c in row["cells"]]
        score = sum(1 for v in row_norms if v in expected_norm)
        if score > best_score:
            best_score = score
            best = idx

    if best_score >= 2:
        return best
    return None


def find_table_header_rows(rows_raw, expected_headers, min_matches=2):
    expected_norm = set(normalize_text(x) for x in expected_headers.keys())
    header_rows = []

    for idx, row in enumerate(rows_raw):
        row_norms = [normalize_text(clean_scalar_text(c["value"])) for c in row["cells"]]
        score = sum(1 for v in row_norms if v in expected_norm)
        if score >= min_matches:
            header_rows.append(idx)

    return header_rows


def build_column_map(header_cells, header_dict):
    col_map = {}
    for cell in header_cells:
        norm = normalize_text(clean_scalar_text(cell["value"]))
        if norm in header_dict:
            col_map[cell["col_idx"]] = header_dict[norm]
    return col_map


def split_rows_from_marker(rows, marker_text):
    marker_norm = normalize_text(marker_text)
    split_idx = None

    for idx, row in enumerate(rows):
        desc_norm = normalize_text(row.get("descricao"))
        if desc_norm == marker_norm:
            split_idx = idx
            break

    if split_idx is None:
        return rows, []

    return rows[:split_idx], rows[split_idx:]


def categorize_section_label(section_label):
    norm_label = normalize_text(section_label)
    if not norm_label:
        return None

    if "RETIRAR" in norm_label:
        return "retirada_cadastro"
    if "SUBSTITUICAO" in norm_label:
        return "substituicao_liberacao"
    if "SUGESTAO" in norm_label and "CADASTR" in norm_label:
        return "sugestao_cadastro"
    if "SUGESTAO" in norm_label and "LIBERAC" in norm_label:
        return "sugestao_liberacao"
    if "OPORTUNIDADE" in norm_label and "CADASTR" in norm_label:
        return "oportunidade_cadastro"
    if "LIBERACAO SCANNTECH" in norm_label or "SCANNTECH" in norm_label:
        return "liberacao_scanntech"
    if "OBJETIVO COMPRA" in norm_label:
        return "objetivo_compra"
    if "LIBERAC" in norm_label:
        return "liberacao"
    if "CADASTR" in norm_label:
        return "cadastro"

    return "secao_especial"


#ajuste aqui ler o bloco sibstituição liberacao
def parse_grid_table(rows_raw, header, sheet_name, form_index, tipo_registro, header_dict, section_rules=False):
    rows = []
    ctx = build_context(header, sheet_name, form_index)
    header_row_idx = detect_table_header_row(rows_raw, header_dict)
    if header_row_idx is None:
        return rows

    table_header_cells = rows_raw[header_row_idx]["cells"]
    col_map = build_column_map(table_header_cells, header_dict)

    normalized_header_labels = set(normalize_text(k) for k in header_dict.keys())

    current_section = None

    # NOVO: tenta localizar uma seção imediatamente acima do cabeçalho
    if section_rules:
        for back_idx in range(header_row_idx - 1, -1, -1):
            back_row = rows_raw[back_idx]
            back_values = [clean_scalar_text(c["value"]) for c in back_row["cells"]]

            if len(back_values) == 1:
                candidate = normalize_text(back_values[0]).rstrip(":")
                if candidate not in normalized_header_labels:
                    current_section = back_values[0].rstrip(":").strip()
                    break

            # se encontrou uma linha com várias células, para de procurar acima
            if len(back_values) > 1:
                break

    line_order = 0

    for idx in range(header_row_idx + 1, len(rows_raw)):
        row = rows_raw[idx]
        values = [clean_scalar_text(c["value"]) for c in row["cells"]]
        norms = [normalize_text(v) for v in values]

        if section_rules:
            if len(values) == 1 and normalize_text(values[0]).rstrip(":") not in normalized_header_labels:
                current_section = values[0].rstrip(":").strip()
                continue

        if sum(1 for n in norms if n in normalized_header_labels) >= 2:
            col_map = build_column_map(row["cells"], header_dict)
            continue

        line_data = {
            **ctx,
            "tipo_registro": tipo_registro,
            "linha_ordem": line_order + 1,
            "row_excel": row["row_excel"],
            "linha_original": row_to_joined(row)
        }

        if current_section:
            line_data["secao_interna"] = current_section
            line_data["categoria_secao"] = categorize_section_label(current_section)

        filled_count = 0
        for cell in row["cells"]:
            col_idx = cell["col_idx"]
            if col_idx in col_map:
                field = col_map[col_idx]
                value = clean_scalar_text(cell["value"])
                if value != "":
                    line_data[field] = value
                    filled_count += 1

        if filled_count >= 1:
            if "quantidade" in line_data:
                line_data["quantidade_numerica"] = parse_number(line_data.get("quantidade"))
            if "volume_mensal" in line_data:
                line_data["volume_mensal_numerica"] = parse_number(line_data.get("volume_mensal"))
            if "volume_periodo" in line_data:
                line_data["volume_periodo_numerica"] = parse_number(line_data.get("volume_periodo"))
            line_order += 1
            line_data["linha_ordem"] = line_order
            rows.append(remove_empty_fields(line_data))

    return rows

def find_month_header_sections(rows_raw):
    sections = []

    for idx, row in enumerate(rows_raw):
        months = []
        for cell in row["cells"]:
            norm = normalize_text(clean_scalar_text(cell["value"]))
            if norm in MONTH_ALIASES:
                months.append((cell["col_idx"], clean_scalar_text(cell["value"]), MONTH_ALIASES[norm]))

        if len(months) >= 3:
            sections.append({
                "header_idx": idx,
                "month_cols": months
            })

    return sections


def parse_month_grid(rows_raw, header, sheet_name, form_index, tipo_registro="encarte_obrigatorio"):
    rows = []
    ctx = build_context(header, sheet_name, form_index)

    sections = find_month_header_sections(rows_raw)
    if not sections:
        return rows

    global_line_visual = 0

    for s_idx, section in enumerate(sections):
        start_idx = section["header_idx"] + 1
        end_idx = len(rows_raw) - 1
        if s_idx < len(sections) - 1:
            end_idx = sections[s_idx + 1]["header_idx"] - 1

        month_cols = section["month_cols"]

        for idx in range(start_idx, end_idx + 1):
            row = rows_raw[idx]
            cell_by_col = {c["col_idx"]: c["value"] for c in row["cells"]}

            added_any = False
            for col_idx, month_label, month_num in month_cols:
                cell_value = cell_by_col.get(col_idx)
                if not cell_value:
                    continue

                products = split_products_from_cell(cell_value)
                for prod in products:
                    item = {
                        **ctx,
                        "tipo_registro": tipo_registro,
                        "mes": month_label,
                        "mes_numero": month_num,
                        "produto": prod,
                        "linha_visual": global_line_visual + 1,
                        "row_excel": row["row_excel"],
                        "linha_original": row_to_joined(row)
                    }
                    rows.append(remove_empty_fields(item))
                    added_any = True

            if added_any:
                global_line_visual += 1

    return rows


def parse_grid_table_multiple_headers(rows_raw, header, sheet_name, form_index, tipo_registro, header_dict, section_rules=False):
    header_rows = find_table_header_rows(rows_raw, header_dict)
    if not header_rows:
        return []

    combined = []
    for order, header_row_idx in enumerate(header_rows):
        start_idx = header_row_idx
        if section_rules and header_row_idx > 0:
            prev_values = [clean_scalar_text(c["value"]) for c in rows_raw[header_row_idx - 1]["cells"]]
            prev_values = [v for v in prev_values if v]
            if len(prev_values) == 1:
                start_idx = header_row_idx - 1

        end_idx = len(rows_raw)
        if order < len(header_rows) - 1:
            end_idx = header_rows[order + 1]

        parsed = parse_grid_table(
            rows_raw[start_idx:end_idx],
            header,
            sheet_name,
            form_index,
            tipo_registro,
            header_dict,
            section_rules=section_rules
        )
        combined.extend(parsed)

    return combined


def slugify_column_name(label):
    norm = normalize_text(label)
    norm = re.sub(r"[^A-Z0-9]+", "_", norm).strip("_")
    return norm.lower() or "coluna"


def parse_objectives_table(rows_raw, header, sheet_name, form_index, title_year=None):
    ctx = build_context(header, sheet_name, form_index)
    header_idx = None

    for idx, row in enumerate(rows_raw):
        values = [clean_scalar_text(c["value"]) for c in row["cells"] if clean_scalar_text(c["value"])]
        if len(values) >= 3 and "OBJETIVOS" not in normalize_text(" | ".join(values)):
            header_idx = idx
            break

    if header_idx is None:
        return []

    header_cells = rows_raw[header_idx]["cells"]
    columns = []
    seen = {}
    for cell in header_cells:
        label = clean_scalar_text(cell["value"])
        if not label:
            continue
        key = slugify_column_name(label)
        seen[key] = seen.get(key, 0) + 1
        if seen[key] > 1:
            key = f"{key}_{seen[key]}"
        columns.append((cell["col_idx"], key))

    if len(columns) < 3:
        return []

    rows = []
    line_order = 0
    for idx in range(header_idx + 1, len(rows_raw)):
        row = rows_raw[idx]
        cell_by_col = {c["col_idx"]: clean_scalar_text(c["value"]) for c in row["cells"]}

        item = {
            **ctx,
            "tipo_registro": "objetivo_tabela",
            "ano_bloco": title_year,
            "row_excel": row["row_excel"],
            "linha_original": row_to_joined(row),
        }

        filled = 0
        for col_idx, key in columns:
            value = cell_by_col.get(col_idx, "")
            if value != "":
                item[key] = value
                filled += 1

        if filled >= 2:
            line_order += 1
            item["linha_ordem"] = line_order
            rows.append(remove_empty_fields(item))

    return rows


def cluster_row_cells(cells, min_gap=4):
    ordered = sorted(cells, key=lambda c: c["col_idx"])
    clusters = []

    for cell in ordered:
        if not clusters:
            clusters.append([cell])
            continue

        if cell["col_idx"] - clusters[-1][-1]["col_idx"] > min_gap:
            clusters.append([cell])
        else:
            clusters[-1].append(cell)

    return clusters


def parse_compact_side_pairs(rows_raw, header, sheet_name, form_index, tipo_registro, title_year=None, ignored_titles=None):
    rows = []
    ctx = build_context(header, sheet_name, form_index)
    ignored_titles = {normalize_text(x) for x in (ignored_titles or [])}
    last_row_by_side = {}
    line_order = 0

    for row in rows_raw:
        clusters = cluster_row_cells(row["cells"])
        for cluster_index, cluster in enumerate(clusters[:2]):
            values = [clean_scalar_text(c["value"]) for c in cluster if clean_scalar_text(c["value"])]
            if not values:
                continue

            first_norm = normalize_text(values[0])
            if first_norm in ignored_titles:
                continue

            side = "esquerda" if cluster_index == 0 else "direita"

            if len(values) >= 2 and not values[1].strip().endswith(":"):
                item = {
                    **ctx,
                    "tipo_registro": tipo_registro,
                    "lado_layout": side,
                    "descricao": values[0],
                    "valor_original": values[1],
                    "valor_numerico": parse_number(values[1]),
                    "unidade": detect_unit(f"{values[0]} {values[1]}"),
                    "ano_bloco": title_year,
                    "row_excel": row["row_excel"],
                    "linha_original": row_to_joined(row),
                }
                if len(values) > 2:
                    item["observacao"] = " | ".join(values[2:])
                line_order += 1
                item["linha_ordem"] = line_order
                rows.append(remove_empty_fields(item))
                last_row_by_side[side] = rows[-1]
            elif len(values) == 1 and side in last_row_by_side and first_norm not in ignored_titles:
                existing_note = last_row_by_side[side].get("observacao")
                note = values[0]
                last_row_by_side[side]["observacao"] = f"{existing_note} | {note}" if existing_note else note

    return rows


def parse_stok_objetivo_compra_table(rows_raw, header, sheet_name, form_index):
    ctx = build_context(header, sheet_name, form_index)
    anchor_idx = None

    for idx, row in enumerate(rows_raw):
        joined = normalize_text(row_to_joined(row))
        if "OBJETIVO COMPRA" in joined and "GATILHO" in joined:
            anchor_idx = idx
            break

    if anchor_idx is None:
        return []

    header_idx = None
    for idx in range(anchor_idx + 1, len(rows_raw)):
        joined = normalize_text(row_to_joined(rows_raw[idx]))
        if "PRODUTO" in joined and "EMBALAGEM" in joined and "INVESTIMENTO" in joined:
            header_idx = idx
            break

    if header_idx is None:
        return []

    col_aliases = {
        "PRODUTO": "produto",
        "EMBALAGEM": "embalagem",
        "JANEIRO A JUNHO": "objetivo_janeiro_junho",
        "JANEIRO À JUNHO": "objetivo_janeiro_junho",
        "JULHO A DEZEMBRO": "objetivo_julho_dezembro",
        "JULHO À DEZEMBRO": "objetivo_julho_dezembro",
        "PERIODO": "periodo",
        "PERÍODO": "periodo",
        "ATINGIMENTO": "atingimento",
        "INVESTIMENTO %": "investimento_percentual",
    }

    header_cells = rows_raw[header_idx]["cells"]
    col_map = {}
    for cell in header_cells:
        norm = normalize_text(clean_scalar_text(cell["value"]))
        if norm in col_aliases:
            col_map[cell["col_idx"]] = col_aliases[norm]

    rows = []
    line_order = 0
    for idx in range(header_idx + 1, len(rows_raw)):
        row = rows_raw[idx]
        values = [clean_scalar_text(c["value"]) for c in row["cells"] if clean_scalar_text(c["value"])]
        if not values:
            continue

        item = {
            **ctx,
            "tipo_registro": "objetivo_compra_stok",
            "row_excel": row["row_excel"],
            "linha_original": row_to_joined(row),
        }

        filled = 0
        for cell in row["cells"]:
            field = col_map.get(cell["col_idx"])
            if not field:
                continue
            value = clean_scalar_text(cell["value"])
            if value != "":
                item[field] = value
                filled += 1

        if filled >= 3:
            line_order += 1
            item["linha_ordem"] = line_order
            rows.append(remove_empty_fields(item))

    return rows


def parse_objetivo_compra_table(rows_raw, header, sheet_name, form_index):
    ctx = build_context(header, sheet_name, form_index)
    anchor_idx = None

    for idx, row in enumerate(rows_raw):
        joined = normalize_text(row_to_joined(row))
        if "OBJETIVO COMPRA" in joined:
            anchor_idx = idx
            break

    if anchor_idx is None:
        return []

    header_idx = None
    for idx in range(anchor_idx + 1, len(rows_raw)):
        joined = normalize_text(row_to_joined(rows_raw[idx]))
        if "PRODUTO" in joined and "EMBALAGEM" in joined and "OBJETIVO TRIMESTRAL" in joined:
            header_idx = idx
            break

    if header_idx is None:
        return []

    header_map = {
        "PRODUTO": "produto",
        "EMBALAGEM": "embalagem",
        "OBJETIVO TRIMESTRAL": "objetivo_trimestral",
    }
    col_map = build_column_map(rows_raw[header_idx]["cells"], header_map)
    rows = []
    line_order = 0

    for idx in range(header_idx + 1, len(rows_raw)):
        row = rows_raw[idx]
        values = [clean_scalar_text(c["value"]) for c in row["cells"] if clean_scalar_text(c["value"])]
        if not values:
            continue

        if len(values) == 1 and not re.search(r"\d", values[0]):
            break

        item = {
            **ctx,
            "tipo_registro": "objetivo_compra",
            "secao_interna": "OBJETIVO COMPRA",
            "row_excel": row["row_excel"],
            "linha_original": row_to_joined(row),
        }

        filled = 0
        for cell in row["cells"]:
            field = col_map.get(cell["col_idx"])
            if not field:
                continue
            value = clean_scalar_text(cell["value"])
            if value != "":
                item[field] = value
                filled += 1

        if filled >= 2:
            objetivo_num = parse_number(item.get("objetivo_trimestral"))
            if objetivo_num is not None:
                item["objetivo_trimestral_numerico"] = objetivo_num
            line_order += 1
            item["linha_ordem"] = line_order
            rows.append(remove_empty_fields(item))

    return rows


def parse_named_contrapartidas_list(rows_raw, header, sheet_name, form_index, section_labels, tipo_registro):
    ctx = build_context(header, sheet_name, form_index)
    normalized_labels = {normalize_text(label) for label in section_labels}
    start_idx = None

    for idx, row in enumerate(rows_raw):
        joined = normalize_text(row_to_joined(row))
        if joined in normalized_labels:
            start_idx = idx + 1
            break

    if start_idx is None:
        return []

    rows = []
    line_order = 0
    for idx in range(start_idx, len(rows_raw)):
        row = rows_raw[idx]
        values = [clean_scalar_text(c["value"]) for c in row["cells"] if clean_scalar_text(c["value"])]
        if not values:
            continue

        if len(values) == 1 and not values[0].isdigit():
            break

        if len(values) >= 3:
            item = {
                **ctx,
                "tipo_registro": tipo_registro,
                "quantidade": values[0],
                "tipo_acao": values[1],
                "periodo": values[2],
                "row_excel": row["row_excel"],
                "linha_original": row_to_joined(row),
            }
            if len(values) >= 4:
                item["lojas"] = values[3]
            quantidade_num = parse_number(values[0])
            if quantidade_num is not None:
                item["quantidade_numerica"] = quantidade_num
            line_order += 1
            item["linha_ordem"] = line_order
            rows.append(remove_empty_fields(item))

    return rows


def detect_encartes_block_name(label_found):
    norm_label = normalize_text(label_found)
    if "ITENS EM ACAO OBRIGATORIOS" in norm_label:
        return "contrapartidas_itens_acao_obrigatorios"
    if "ITENS PONTA DE GONDOLA OBRIGATORIOS" in norm_label:
        return "contrapartidas_itens_ponta_gondola_obrigatorios"
    if "SUGESTAO" in norm_label:
        return "encartes_sugestao"
    return "encartes_obrigatorios"


def split_contrapartidas_and_encartes(rows_raw):
    split_idx = None
    encartes_block_name = None
    split_labels = {
        "SUGESTAO DE ENCARTES",
        "SUGESTÃO DE ENCARTES",
        "ENCARTES OBRIGATORIOS",
        "ENCARTES OBRIGATÓRIOS",
        "ENCARTES SUGESTAO",
        "ENCARTES SUGESTÃO",
        "ITENS EM ACAO OBRIGATORIOS",
        "ITENS EM AÇÃO OBRIGATÓRIOS",
        "ITENS PONTA DE GONDOLA OBRIGATORIOS",
        "ITENS PONTA DE GÔNDOLA OBRIGATÓRIOS",
    }

    norm_split_labels = {normalize_text(x) for x in split_labels}

    for idx, row in enumerate(rows_raw):
        joined = normalize_text(row_to_joined(row))
        values = [normalize_text(clean_scalar_text(c["value"])) for c in row["cells"]]

        for label in norm_split_labels:
            if joined == label or joined.startswith(label):
                split_idx = idx
                encartes_block_name = detect_encartes_block_name(row_to_joined(row))
                break

        if split_idx is not None:
            break

        if len(values) == 1 and values[0] in norm_split_labels:
            split_idx = idx
            encartes_block_name = detect_encartes_block_name(values[0])
            break

    if split_idx is None:
        return rows_raw, [], None

    contrapartidas_rows = rows_raw[:split_idx]
    encartes_rows = rows_raw[split_idx + 1:]
    return contrapartidas_rows, encartes_rows, encartes_block_name


def infer_section_bucket(section_label, index_order):
    norm_label = normalize_text(section_label)
    if "LIBERAC" in norm_label:
        return "sugestao_liberacao"
    if "CADASTR" in norm_label:
        return "oportunidade_cadastro"
    if index_order == 0:
        return "oportunidade_cadastro"
    return "sugestao_liberacao"


def parse_oportunidades_cadastros_liberacoes(rows_raw, header, sheet_name, form_index):
    oportunidades_rows = []
    sugestao_rows = []

    header_rows = find_table_header_rows(rows_raw, CADASTROS_HEADERS)
    if not header_rows:
        return oportunidades_rows, sugestao_rows

    normalized_header_labels = {normalize_text(k) for k in CADASTROS_HEADERS.keys()}

    for order, header_row_idx in enumerate(header_rows):
        end_idx = len(rows_raw)
        if order < len(header_rows) - 1:
            end_idx = header_rows[order + 1]

        section_label = ""
        for back_idx in range(header_row_idx - 1, -1, -1):
            values = [clean_scalar_text(c["value"]) for c in rows_raw[back_idx]["cells"]]
            values = [v for v in values if v]

            if not values:
                continue

            if len(values) == 1:
                candidate = values[0].rstrip(":").strip()
                if normalize_text(candidate) not in normalized_header_labels:
                    section_label = candidate
                    break

            if len(values) > 1:
                break

        tipo_registro = infer_section_bucket(section_label, order)
        parsed_rows = parse_grid_table(
            rows_raw[header_row_idx:end_idx],
            header,
            sheet_name,
            form_index,
            tipo_registro,
            CADASTROS_HEADERS
        )

        if section_label:
            for item in parsed_rows:
                item["secao_interna"] = section_label
                item["categoria_secao"] = categorize_section_label(section_label)

        if tipo_registro == "sugestao_liberacao":
            sugestao_rows.extend(parsed_rows)
        else:
            oportunidades_rows.extend(parsed_rows)

    return oportunidades_rows, sugestao_rows


def normalize_detected_block_name(block_name, label_found):
    if block_name == "encartes_obrigatorios":
        return detect_encartes_block_name(label_found)
    return block_name


def split_premissas_from_result(result):
    descricao_rows = result.get("descricao_investimento_rows", [])
    investimentos_rows = result.get("investimentos_extras_rows", [])

    descricao_main, descricao_premissas = split_rows_from_marker(descricao_rows, "PREMISSAS GERAIS")
    investimentos_main, investimentos_premissas = split_rows_from_marker(investimentos_rows, "PREMISSAS GERAIS")

    premissas_rows = descricao_premissas + investimentos_premissas
    if not premissas_rows:
        return

    result["descricao_investimento_rows"] = descricao_main
    result["investimentos_extras_rows"] = investimentos_main
    result["premissas_gerais_rows"] = premissas_rows

    if "premissas_gerais" not in result["identified_blocks"]:
        result["identified_blocks"].append("premissas_gerais")


def split_objetivo_compra_from_cadastros(result):
    objetivo_rows = result.get("objetivo_compra_rows", [])
    cadastros_rows = result.get("cadastros_vinculados_rows", [])

    if not objetivo_rows and cadastros_rows:
        objetivo_rows = [row for row in cadastros_rows if row.get("categoria_secao") == "objetivo_compra"]

    if not objetivo_rows:
        return

    result["objetivo_compra_rows"] = objetivo_rows
    result["cadastros_vinculados_rows"] = [
        row for row in cadastros_rows
        if row.get("categoria_secao") != "objetivo_compra"
    ]

    if "objetivo_compra" not in result["identified_blocks"]:
        result["identified_blocks"].append("objetivo_compra")


def process_form_grid(grid, sheet_name, form_index, region_meta):
    headers_found = find_block_headers(grid)
    if not is_valid_form_region(headers_found):
        return None

    ranges = build_row_ranges(headers_found, len(grid))

    result = {
        "sheet_name": sheet_name,
        "formulario_index": form_index,
        "region_start_col": region_meta["start_col"] + 1,
        "region_end_col": region_meta["end_col"] + 1,
        "region_width": region_meta["width"],
        "identified_blocks": [normalize_detected_block_name(h["block"], h["label_found"]) for h in headers_found],
        "header": {
            "cliente": None,
            "periodo_original": None,
            "periodo_inicio": None,
            "periodo_fim": None,
            "periodo_normalizado": None,
            "periodo_status": None,
            "periodo_acoes_original": None,
            "periodo_acoes_inicio": None,
            "periodo_acoes_fim": None,
            "periodo_acoes_normalizado": None,
            "periodo_acoes_status": None,
            "numero_acordo": None,
            "titulo_plano": None
        },
        "plano_negocios_rows": [],
        "historico_rows": [],
        "objetivos_rows": [],
        "objetivos_compactos_rows": [],
        "descricao_investimento_rows": [],
        "descricao_investimento_compacto_rows": [],
        "premissas_gerais_rows": [],
        "contrapartidas_rows": [],
        "contrapartidas_itens_foco_rows": [],
        "contrapartidas_acoes_rows": [],
        "contrapartidas_itens_acao_obrigatorios_rows": [],
        "contrapartidas_itens_ponta_gondola_obrigatorios_rows": [],
        "contrapartidas_encartes_mensal_rows": [],
        "encartes_obrigatorios_rows": [],
        "encartes_sugestao_rows": [],
        "cadastros_vinculados_rows": [],
        "situacao_liberacao_rows": [],
        "oportunidades_cadastros_rows": [],
        "sugestao_liberacao_rows": [],
        "investimentos_extras_rows": [],
        "objetivo_compra_rows": [],
        "stok_objetivo_compra_rows": [],
        "raw_blocks": []
    }

    for block_info in ranges:
        effective_block = normalize_detected_block_name(block_info["block"], block_info["label_found"])
        rows_raw = slice_block_rows(grid, block_info)
        title_year = extract_year_from_title(block_info["label_found"])

        result["raw_blocks"].append({
            "block": effective_block,
            "label_found": clean_scalar_text(block_info["label_found"]),
            "start_row": block_info["start_row"] + 1,
            "end_row": block_info["end_row"] + 1,
            "total_rows_raw": len(rows_raw)
        })

        if effective_block == "plano_negocios":
            result["header"] = extract_header_from_block(rows_raw, block_info["label_found"])
            result["plano_negocios_rows"] = [remove_empty_fields(build_context(result["header"], sheet_name, form_index))]

        elif effective_block == "historico":
            result["historico_rows"].extend(
                parse_kv_list(rows_raw, result["header"], sheet_name, form_index, "historico", title_year)
            )

        elif effective_block == "objetivos":
            objetivos_table_rows = parse_objectives_table(
                rows_raw,
                result["header"],
                sheet_name,
                form_index,
                title_year
            )
            if objetivos_table_rows:
                result["objetivos_rows"].extend(objetivos_table_rows)
            else:
                result["objetivos_rows"].extend(
                    parse_kv_list(rows_raw, result["header"], sheet_name, form_index, "objetivo", title_year)
                )

            if not objetivos_table_rows:
                objetivos_compactos_rows = parse_compact_side_pairs(
                    rows_raw,
                    result["header"],
                    sheet_name,
                    form_index,
                    "objetivo_compacto",
                    title_year,
                    ignored_titles=["OBJETIVOS PARA O INVESTIMENTO", "OBJETIVOS", block_info["label_found"]]
                )
                if objetivos_compactos_rows:
                    result["objetivos_compactos_rows"].extend(objetivos_compactos_rows)
                    if "objetivos_compactos" not in result["identified_blocks"]:
                        result["identified_blocks"].append("objetivos_compactos")

        elif effective_block == "descricao_investimento":
            result["descricao_investimento_rows"].extend(
                parse_kv_list(rows_raw, result["header"], sheet_name, form_index, "descricao_investimento", None)
            )
            descricao_compacta_rows = parse_compact_side_pairs(
                rows_raw,
                result["header"],
                sheet_name,
                form_index,
                "descricao_investimento_compacto",
                None,
                ignored_titles=["DESCRIÇÃO DO INVESTIMENTO", "DESCRICAO DO INVESTIMENTO", block_info["label_found"]]
            )
            if descricao_compacta_rows:
                result["descricao_investimento_compacto_rows"].extend(descricao_compacta_rows)
                if "descricao_investimento_compacto" not in result["identified_blocks"]:
                    result["identified_blocks"].append("descricao_investimento_compacto")

        elif effective_block == "contrapartidas":
            encartes_mensal_rows = parse_named_contrapartidas_list(
                rows_raw,
                result["header"],
                sheet_name,
                form_index,
                ["ENCATES MENSAL", "ENCARTES MENSAL"],
                "contrapartida_encarte_mensal"
            )
            if encartes_mensal_rows:
                result["contrapartidas_encartes_mensal_rows"].extend(encartes_mensal_rows)
                if "contrapartidas_encartes_mensal" not in result["identified_blocks"]:
                    result["identified_blocks"].append("contrapartidas_encartes_mensal")

            contrapartidas_only_rows, encartes_embedded_rows, encartes_block_name = split_contrapartidas_and_encartes(rows_raw)

            result["contrapartidas_rows"].extend(
                parse_grid_table(
                    contrapartidas_only_rows,
                    result["header"],
                    sheet_name,
                    form_index,
                    "contrapartida",
                    CONTRAPARTIDAS_HEADERS
                )
            )

            if encartes_embedded_rows:
                if encartes_block_name == "contrapartidas_itens_acao_obrigatorios":
                    result["contrapartidas_itens_acao_obrigatorios_rows"].extend(
                        parse_month_grid(
                            encartes_embedded_rows,
                            result["header"],
                            sheet_name,
                            form_index,
                            "contrapartida_item_acao_obrigatorio"
                        )
                    )
                elif encartes_block_name == "contrapartidas_itens_ponta_gondola_obrigatorios":
                    result["contrapartidas_itens_ponta_gondola_obrigatorios_rows"].extend(
                        parse_month_grid(
                            encartes_embedded_rows,
                            result["header"],
                            sheet_name,
                            form_index,
                            "contrapartida_item_ponta_gondola_obrigatorio"
                        )
                    )
                else:
                    encartes_key = "encartes_sugestao_rows" if encartes_block_name == "encartes_sugestao" else "encartes_obrigatorios_rows"
                    result[encartes_key].extend(
                        parse_month_grid(encartes_embedded_rows, result["header"], sheet_name, form_index)
                    )

                if encartes_block_name and encartes_block_name not in result["identified_blocks"]:
                    result["identified_blocks"].append(encartes_block_name)

        elif effective_block == "contrapartidas_itens_foco":
            result["contrapartidas_itens_foco_rows"].extend(
                parse_grid_table(
                    rows_raw,
                    result["header"],
                    sheet_name,
                    form_index,
                    "contrapartida_item_foco",
                    ITENS_FOCO_HEADERS
                )
            )

        elif effective_block == "contrapartidas_acoes":
            contrapartidas_only_rows, encartes_embedded_rows, encartes_block_name = split_contrapartidas_and_encartes(rows_raw)

            result["contrapartidas_acoes_rows"].extend(
                parse_grid_table(
                    contrapartidas_only_rows,
                    result["header"],
                    sheet_name,
                    form_index,
                    "contrapartida_acao",
                    CONTRAPARTIDAS_HEADERS
                )
            )

            if encartes_embedded_rows:
                if encartes_block_name == "contrapartidas_itens_acao_obrigatorios":
                    result["contrapartidas_itens_acao_obrigatorios_rows"].extend(
                        parse_month_grid(
                            encartes_embedded_rows,
                            result["header"],
                            sheet_name,
                            form_index,
                            "contrapartida_item_acao_obrigatorio"
                        )
                    )
                elif encartes_block_name == "contrapartidas_itens_ponta_gondola_obrigatorios":
                    result["contrapartidas_itens_ponta_gondola_obrigatorios_rows"].extend(
                        parse_month_grid(
                            encartes_embedded_rows,
                            result["header"],
                            sheet_name,
                            form_index,
                            "contrapartida_item_ponta_gondola_obrigatorio"
                        )
                    )
                else:
                    encartes_key = "encartes_sugestao_rows" if encartes_block_name == "encartes_sugestao" else "encartes_obrigatorios_rows"
                    result[encartes_key].extend(
                        parse_month_grid(encartes_embedded_rows, result["header"], sheet_name, form_index)
                    )

                if encartes_block_name and encartes_block_name not in result["identified_blocks"]:
                    result["identified_blocks"].append(encartes_block_name)

        elif effective_block == "encartes_obrigatorios":
            result["encartes_obrigatorios_rows"].extend(
                parse_month_grid(rows_raw, result["header"], sheet_name, form_index)
            )

        elif effective_block == "encartes_sugestao":
            result["encartes_sugestao_rows"].extend(
                parse_month_grid(rows_raw, result["header"], sheet_name, form_index)
            )

        elif effective_block == "cadastros_vinculados":
            result["cadastros_vinculados_rows"].extend(
                parse_grid_table_multiple_headers(
                    rows_raw,
                    result["header"],
                    sheet_name,
                    form_index,
                    "cadastro_vinculado",
                    CADASTROS_HEADERS,
                    section_rules=True
                )
            )
            objetivo_compra_rows = parse_objetivo_compra_table(
                rows_raw,
                result["header"],
                sheet_name,
                form_index
            )
            if objetivo_compra_rows:
                result["objetivo_compra_rows"].extend(objetivo_compra_rows)
                if "objetivo_compra" not in result["identified_blocks"]:
                    result["identified_blocks"].append("objetivo_compra")

        elif effective_block == "oportunidades_cadastros_liberacoes":
            oportunidades_rows, sugestao_rows = parse_oportunidades_cadastros_liberacoes(
                rows_raw,
                result["header"],
                sheet_name,
                form_index
            )
            result["oportunidades_cadastros_rows"].extend(oportunidades_rows)
            result["sugestao_liberacao_rows"].extend(sugestao_rows)

        elif effective_block == "situacao_liberacao":
            result["situacao_liberacao_rows"].extend(
                parse_grid_table_multiple_headers(
                    rows_raw,
                    result["header"],
                    sheet_name,
                    form_index,
                    "situacao_liberacao",
                    CADASTROS_HEADERS,
                    section_rules=True
                )
            )

        elif effective_block == "investimentos_extras":
            result["investimentos_extras_rows"].extend(
                parse_kv_list(rows_raw, result["header"], sheet_name, form_index, "investimento_extra", None)
            )
            stok_rows = parse_stok_objetivo_compra_table(
                rows_raw,
                result["header"],
                sheet_name,
                form_index
            )
            if stok_rows:
                result["stok_objetivo_compra_rows"].extend(stok_rows)
                if "stok_objetivo_compra" not in result["identified_blocks"]:
                    result["identified_blocks"].append("stok_objetivo_compra")

    if not result["plano_negocios_rows"]:
        result["plano_negocios_rows"] = [remove_empty_fields(build_context(result["header"], sheet_name, form_index))]

    split_premissas_from_result(result)
    split_objetivo_compra_from_cadastros(result)
    result["header"] = remove_empty_fields(result["header"])
    result["identified_blocks"] = list(dict.fromkeys(result["identified_blocks"]))
    return remove_empty_fields(result)


def build_manual_compact_record(form, sheet_name, row_excel, descricao, valor, tipo_registro, observacao=None, lado_layout="direita"):
    item = {
        **build_context(form.get("header", {}), sheet_name, form.get("formulario_index")),
        "tipo_registro": tipo_registro,
        "lado_layout": lado_layout,
        "descricao": clean_scalar_text(descricao),
        "valor_original": clean_scalar_text(valor),
        "valor_numerico": parse_number(valor),
        "unidade": detect_unit(f"{descricao} {valor}"),
        "row_excel": row_excel,
        "linha_original": " | ".join(x for x in [clean_scalar_text(descricao), clean_scalar_text(valor), clean_scalar_text(observacao)] if x),
    }
    if observacao:
        item["observacao"] = clean_scalar_text(observacao)
    return remove_empty_fields(item)


def append_unique_rows(target_rows, new_rows):
    existing = {
        (
            row.get("tipo_registro"),
            row.get("descricao"),
            row.get("valor_original"),
            row.get("observacao"),
        )
        for row in target_rows
    }

    for row in new_rows:
        key = (
            row.get("tipo_registro"),
            row.get("descricao"),
            row.get("valor_original"),
            row.get("observacao"),
        )
        if key not in existing:
            row["linha_ordem"] = len(target_rows) + 1
            target_rows.append(row)
            existing.add(key)


def enrich_bonato_second_form(ws, forms):
    if ws.title != "BONATO":
        return

    for form in forms:
        if form.get("formulario_index") != 2:
            continue

        objetivos_rows = [
            build_manual_compact_record(form, ws.title, 16, "OBJETIVO DE VOLUME/MÊS (CAIXAS):", ws.cell(16, 42).value, "objetivo_compacto"),
            build_manual_compact_record(form, ws.title, 19, "OBJETIVO DE VOLUME/TOTAL DO PERÍODO (CAIXAS):", ws.cell(19, 42).value, "objetivo_compacto"),
            build_manual_compact_record(form, ws.title, 22, "MODALIDADE DE CARREGAMENTO:", ws.cell(22, 42).value, "objetivo_compacto"),
        ]
        objetivos_rows = [row for row in objetivos_rows if row.get("valor_original")]
        if objetivos_rows:
            append_unique_rows(form.setdefault("objetivos_compactos_rows", []), objetivos_rows)
            if "objetivos_compactos" not in form.setdefault("identified_blocks", []):
                form["identified_blocks"].append("objetivos_compactos")

        descricao_rows = [
            build_manual_compact_record(form, ws.title, 27, "VALOR DE INVESTIMENTO:", ws.cell(27, 42).value, "descricao_investimento_compacto"),
            build_manual_compact_record(
                form,
                ws.title,
                30,
                "FORMA DE PAGAMENTO:",
                ws.cell(30, 42).value,
                "descricao_investimento_compacto",
                observacao=ws.cell(30, 47).value,
            ),
        ]
        descricao_rows = [row for row in descricao_rows if row.get("valor_original")]
        if descricao_rows:
            append_unique_rows(form.setdefault("descricao_investimento_compacto_rows", []), descricao_rows)
            if "descricao_investimento_compacto" not in form.setdefault("identified_blocks", []):
                form["identified_blocks"].append("descricao_investimento_compacto")


def process_sheet(ws):
    full_grid = worksheet_to_grid(ws)

    anchors = find_form_anchors(full_grid)
    if len(anchors) >= 1:
        regions = detect_form_regions_from_anchors(full_grid, anchors)
    else:
        regions = detect_column_regions(full_grid)

    forms = []
    ignored_regions = []

    form_index = 1
    for region in regions:
        region_grid = crop_grid_to_region(full_grid, region["start_col"], region["end_col"])
        headers_found = find_block_headers(region_grid)

        if is_valid_form_region(headers_found):
            processed = process_form_grid(region_grid, ws.title, form_index, region)
            if processed:
                forms.append(processed)
                form_index += 1
        else:
            ignored_regions.append({
                "start_col": region["start_col"] + 1,
                "end_col": region["end_col"] + 1,
                "width": region["width"],
                "motivo": "regiao_sem_blocos_principais"
            })

    enrich_bonato_second_form(ws, forms)
    return remove_empty_fields({
        "sheet_name": ws.title,
        "forms": forms,
        "ignored_regions": ignored_regions
    })


def main():
    try:
        if len(sys.argv) < 2:
            raise Exception("Caminho do arquivo não informado.")

        file_path = sys.argv[1]
        if not os.path.exists(file_path):
            raise Exception("Arquivo não encontrado.")

        wb = load_workbook(file_path, data_only=True)

        output = {
            "success": True,
            "file_name": os.path.basename(file_path),
            "processed_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_sheets": len(wb.sheetnames),
            "sheets": []
        }

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            output["sheets"].append(process_sheet(ws))

        print(json.dumps(remove_empty_fields(output), ensure_ascii=False))

    except Exception as e:
        print(json.dumps({
            "success": False,
            "error": str(e)
        }, ensure_ascii=False))


if __name__ == "__main__":
    main()
